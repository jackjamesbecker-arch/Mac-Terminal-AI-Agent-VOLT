#!/usr/bin/env python3
"""
volt - A terminal AI agent powered by Groq
Interactive REPL with file, shell, web search, notes, clipboard, weather, and more.
"""

import json
import os
import re
import subprocess
import sys
import time
import urllib.request
import urllib.parse
import urllib.error
from datetime import datetime
from pathlib import Path

# ── Config ────────────────────────────────────────────────────────────────────

MODEL_FAST  = "meta-llama/llama-4-scout-17b-16e-instruct"
MODEL_SMART = "meta-llama/llama-4-scout-17b-16e-instruct"
MODEL       = MODEL_FAST
MAX_TOKENS  = 4096

VOLT_DIR    = Path.home() / ".volt"
NOTES_FILE  = VOLT_DIR / "notes.json"
HISTORY_DIR = VOLT_DIR / "history"
MEMORY_FILE = VOLT_DIR / "memory.json"
NOVA_SYNC   = VOLT_DIR / "nova_sync.json"
NOVA_URL    = "https://nova-ai-weld-nu.vercel.app/?beta=APEX0777"

# ── Messaging Credentials ─────────────────────────────────────────────────────
# Email — use a Gmail App Password (not your real password)
#   Setup: myaccount.google.com → Security → 2-Step Verification → App passwords
EMAIL_ADDRESS  = os.environ.get("VOLT_EMAIL", "")
EMAIL_PASSWORD = os.environ.get("VOLT_EMAIL_PASSWORD", "")

# ── User Profiles ─────────────────────────────────────────────────────────────

USERS = {
    "7474":    {"name": "Jackson", "color": "\033[96m",  "emoji": "⚡"},  # cyan
    "110583":  {"name": "Jules",   "color": "\033[95m",  "emoji": "✨"},  # purple
    "8675309": {"name": "Larry",   "color": "\033[92m",  "emoji": "👑"},  # green
}

QUOTES = [
    "The best way to predict the future is to build it.",
    "Move fast and break things. Then fix them.",
    "Code is like humor. When you have to explain it, it's bad.",
    "First, solve the problem. Then, write the code.",
    "Simplicity is the soul of efficiency.",
    "Make it work, make it right, make it fast.",
    "Any fool can write code that a computer can understand.",
]

# ── ANSI Colors ───────────────────────────────────────────────────────────────

RESET  = "\033[0m"
BOLD   = "\033[1m"
DIM    = "\033[2m"
CYAN   = "\033[96m"
GREEN  = "\033[92m"
YELLOW = "\033[93m"
RED    = "\033[91m"
PURPLE = "\033[95m"
BLUE   = "\033[94m"

def c(color: str, text: str) -> str:
    return f"{color}{text}{RESET}"

# ── Tool Definitions ──────────────────────────────────────────────────────────

TOOLS = [
    {"type": "function", "function": {
        "name": "read_file",
        "description": "Read the contents of a file.",
        "parameters": {"type": "object", "properties": {
            "path": {"type": "string", "description": "Path to the file"}
        }, "required": ["path"]}
    }},
    {"type": "function", "function": {
        "name": "write_file",
        "description": "Write or overwrite a file. Creates parent directories if needed.",
        "parameters": {"type": "object", "properties": {
            "path":    {"type": "string", "description": "Path to the file"},
            "content": {"type": "string", "description": "Content to write"}
        }, "required": ["path", "content"]}
    }},
    {"type": "function", "function": {
        "name": "run_command",
        "description": "Execute a shell command and return output.",
        "parameters": {"type": "object", "properties": {
            "command": {"type": "string", "description": "Shell command to run"},
            "cwd":     {"type": "string", "description": "Working directory (optional)"}
        }, "required": ["command"]}
    }},
    {"type": "function", "function": {
        "name": "web_search",
        "description": "Search the web for current information.",
        "parameters": {"type": "object", "properties": {
            "query": {"type": "string", "description": "Search query"}
        }, "required": ["query"]}
    }},
    {"type": "function", "function": {
        "name": "list_dir",
        "description": "List files and directories at a path.",
        "parameters": {"type": "object", "properties": {
            "path": {"type": "string", "description": "Directory path (defaults to current)"}
        }, "required": []}
    }},
    {"type": "function", "function": {
        "name": "save_note",
        "description": "Save a named note to persistent storage.",
        "parameters": {"type": "object", "properties": {
            "name":    {"type": "string", "description": "Note name/title"},
            "content": {"type": "string", "description": "Note content"}
        }, "required": ["name", "content"]}
    }},
    {"type": "function", "function": {
        "name": "get_note",
        "description": "Retrieve a saved note by name. Use list_notes to see all notes.",
        "parameters": {"type": "object", "properties": {
            "name": {"type": "string", "description": "Note name to retrieve"}
        }, "required": ["name"]}
    }},
    {"type": "function", "function": {
        "name": "list_notes",
        "description": "List all saved notes.",
        "parameters": {"type": "object", "properties": {}, "required": []}
    }},
    {"type": "function", "function": {
        "name": "clipboard_copy",
        "description": "Copy text to the clipboard.",
        "parameters": {"type": "object", "properties": {
            "text": {"type": "string", "description": "Text to copy"}
        }, "required": ["text"]}
    }},
    {"type": "function", "function": {
        "name": "clipboard_paste",
        "description": "Read the current clipboard contents.",
        "parameters": {"type": "object", "properties": {}, "required": []}
    }},
    {"type": "function", "function": {
        "name": "get_weather",
        "description": "Get current weather for a location.",
        "parameters": {"type": "object", "properties": {
            "location": {"type": "string", "description": "City name or 'here' for current location"}
        }, "required": ["location"]}
    }},
    {"type": "function", "function": {
        "name": "remember",
        "description": "Save something to memory to recall later across sessions.",
        "parameters": {"type": "object", "properties": {
            "key":   {"type": "string", "description": "Memory key/label"},
            "value": {"type": "string", "description": "Value to remember"}
        }, "required": ["key", "value"]}
    }},
    {"type": "function", "function": {
        "name": "recall",
        "description": "Recall something from memory.",
        "parameters": {"type": "object", "properties": {
            "key": {"type": "string", "description": "Memory key to look up (or 'all' for everything)"}
        }, "required": ["key"]}
    }},
    {"type": "function", "function": {
        "name": "notify",
        "description": "Send a Mac desktop notification.",
        "parameters": {"type": "object", "properties": {
            "title":   {"type": "string", "description": "Notification title"},
            "message": {"type": "string", "description": "Notification body"}
        }, "required": ["title", "message"]}
    }},
    {"type": "function", "function": {
        "name": "send_imessage",
        "description": "Send an iMessage or SMS to a phone number or Apple ID email.",
        "parameters": {"type": "object", "properties": {
            "to":      {"type": "string", "description": "Phone number (e.g. +15551234567) or Apple ID email"},
            "message": {"type": "string", "description": "Message text to send"}
        }, "required": ["to", "message"]}
    }},
    {"type": "function", "function": {
        "name": "send_email",
        "description": "Send an email via Gmail, optionally with a file attachment.",
        "parameters": {"type": "object", "properties": {
            "to":         {"type": "string", "description": "Recipient email address"},
            "subject":    {"type": "string", "description": "Email subject line"},
            "body":       {"type": "string", "description": "Email body text"},
            "attachment": {"type": "string", "description": "Path to file to attach (optional)"}
        }, "required": ["to", "subject", "body"]}
    }},
    {"type": "function", "function": {
        "name": "receive_files",
        "description": "Check Gmail inbox for emails with attachments and download them.",
        "parameters": {"type": "object", "properties": {
            "save_dir": {"type": "string", "description": "Where to save files (defaults to ~/Downloads/volt_received)"},
            "limit":    {"type": "integer", "description": "How many recent emails to check (default 5)"}
        }, "required": []}
    }},
    {"type": "function", "function": {
        "name": "lookup_contact",
        "description": "Look up a contact by name from Mac Contacts app. Returns their phone number and email.",
        "parameters": {"type": "object", "properties": {
            "name": {"type": "string", "description": "Contact name to search for"}
        }, "required": ["name"]}
    }},
    {"type": "function", "function": {
        "name": "message_contact",
        "description": "Send an iMessage to a contact by name. Looks them up automatically.",
        "parameters": {"type": "object", "properties": {
            "name":    {"type": "string", "description": "Contact name (e.g. 'Jules')"},
            "message": {"type": "string", "description": "Message to send"}
        }, "required": ["name", "message"]}
    }},
    {"type": "function", "function": {
        "name": "email_contact",
        "description": "Send an email to a contact by name. Looks them up automatically.",
        "parameters": {"type": "object", "properties": {
            "name":    {"type": "string", "description": "Contact name (e.g. 'Mom')"},
            "subject": {"type": "string", "description": "Email subject"},
            "body":    {"type": "string", "description": "Email body"}
        }, "required": ["name", "subject", "body"]}
    }},
    {"type": "function", "function": {
        "name": "get_calendar",
        "description": "Get calendar events for a time period.",
        "parameters": {"type": "object", "properties": {
            "period": {"type": "string", "description": "Time period: 'today', 'tomorrow', 'this week', or a date like '2024-12-25'"}
        }, "required": ["period"]}
    }},
    {"type": "function", "function": {
        "name": "add_reminder",
        "description": "Add a reminder to Mac Reminders app.",
        "parameters": {"type": "object", "properties": {
            "title": {"type": "string", "description": "Reminder title"},
            "notes": {"type": "string", "description": "Additional notes (optional)"},
            "due":   {"type": "string", "description": "Due date/time e.g. 'tomorrow at 9am', 'in 30 minutes' (optional)"}
        }, "required": ["title"]}
    }},
    {"type": "function", "function": {
        "name": "take_screenshot",
        "description": "Take a screenshot of the screen and save it.",
        "parameters": {"type": "object", "properties": {
            "path": {"type": "string", "description": "Where to save it (optional, defaults to Desktop)"}
        }, "required": []}
    }},
    {"type": "function", "function": {
        "name": "open_url",
        "description": "Open a URL in the default browser.",
        "parameters": {"type": "object", "properties": {
            "url": {"type": "string", "description": "URL to open"}
        }, "required": ["url"]}
    }},
    {"type": "function", "function": {
        "name": "apple_music",
        "description": "Control Apple Music: play, pause, skip, previous, volume, search for a song/artist, or see what's playing.",
        "parameters": {"type": "object", "properties": {
            "action": {"type": "string", "description": "Action: play, pause, next, previous, volume, search, current"},
            "query":  {"type": "string", "description": "Song or artist name for search, or volume level 0-100 for volume (optional)"}
        }, "required": ["action"]}
    }},
    {"type": "function", "function": {
        "name": "read_csv",
        "description": "Read and summarize a CSV or Excel file.",
        "parameters": {"type": "object", "properties": {
            "path": {"type": "string", "description": "Path to the CSV or Excel file"}
        }, "required": ["path"]}
    }},
    {"type": "function", "function": {
        "name": "schedule_task",
        "description": "Schedule a recurring task. Lists or removes existing tasks.",
        "parameters": {"type": "object", "properties": {
            "action":  {"type": "string", "description": "Action: add, list, remove"},
            "task":    {"type": "string", "description": "Task description (for add)"},
            "schedule":{"type": "string", "description": "When to run e.g. 'every day at 9am' (for add)"},
            "task_id": {"type": "string", "description": "Task ID to remove (for remove)"}
        }, "required": ["action"]}
    }},
    {"type": "function", "function": {
        "name": "daily_briefing",
        "description": "Give a morning briefing: weather, calendar events, and top news.",
        "parameters": {"type": "object", "properties": {
            "location": {"type": "string", "description": "City for weather (optional, defaults to current location)"}
        }, "required": []}
    }},
    {"type": "function", "function": {
        "name": "summarize_url",
        "description": "Fetch and summarize the content of a webpage or article by URL.",
        "parameters": {"type": "object", "properties": {
            "url": {"type": "string", "description": "URL to summarize"}
        }, "required": ["url"]}
    }},
    {"type": "function", "function": {
        "name": "get_active_app",
        "description": "Get the name of the currently active/focused app on Mac.",
        "parameters": {"type": "object", "properties": {}, "required": []}
    }},
    {"type": "function", "function": {
        "name": "check_website",
        "description": "Check if a website is up or down.",
        "parameters": {"type": "object", "properties": {
            "url": {"type": "string", "description": "Website URL to check"}
        }, "required": ["url"]}
    }},
    {"type": "function", "function": {
        "name": "generate_password",
        "description": "Generate a secure random password and optionally save it.",
        "parameters": {"type": "object", "properties": {
            "length": {"type": "integer", "description": "Password length (default 20)"},
            "label":  {"type": "string", "description": "Label to save it under in memory (optional)"}
        }, "required": []}
    }},
    {"type": "function", "function": {
        "name": "get_stock",
        "description": "Get current stock or crypto price.",
        "parameters": {"type": "object", "properties": {
            "symbol": {"type": "string", "description": "Stock ticker or crypto symbol e.g. AAPL, BTC, ETH"}
        }, "required": ["symbol"]}
    }},
    {"type": "function", "function": {
        "name": "translate",
        "description": "Translate text to another language.",
        "parameters": {"type": "object", "properties": {
            "text":     {"type": "string", "description": "Text to translate"},
            "language": {"type": "string", "description": "Target language e.g. Spanish, French, Japanese"}
        }, "required": ["text", "language"]}
    }},
    {"type": "function", "function": {
        "name": "nova_open",
        "description": "Open NOVA in the browser.",
        "parameters": {"type": "object", "properties": {}, "required": []}
    }},
    {"type": "function", "function": {
        "name": "nova_status",
        "description": "Check if NOVA is online and get its current status.",
        "parameters": {"type": "object", "properties": {}, "required": []}
    }},
    {"type": "function", "function": {
        "name": "nova_sync",
        "description": "Sync a key-value to NOVA's shared memory file so NOVA can read it.",
        "parameters": {"type": "object", "properties": {
            "key":   {"type": "string", "description": "Key to sync"},
            "value": {"type": "string", "description": "Value to sync"}
        }, "required": ["key", "value"]}
    }},
    {"type": "function", "function": {
        "name": "todo",
        "description": "Manage a to-do list. Add, complete, remove, or list tasks.",
        "parameters": {"type": "object", "properties": {
            "action": {"type": "string", "description": "Action: add, done, remove, list, clear"},
            "task":   {"type": "string", "description": "Task text (for add/done/remove)"},
            "id":     {"type": "integer", "description": "Task ID number (for done/remove)"}
        }, "required": ["action"]}
    }},
    {"type": "function", "function": {
        "name": "set_timer",
        "description": "Set a countdown timer. Sends a Mac notification when done.",
        "parameters": {"type": "object", "properties": {
            "duration": {"type": "string", "description": "Duration e.g. '25 minutes', '1 hour', '30 seconds'"}
        }, "required": ["duration"]}
    }},
    {"type": "function", "function": {
        "name": "get_battery",
        "description": "Get Mac battery level and charging status.",
        "parameters": {"type": "object", "properties": {}, "required": []}
    }},
    {"type": "function", "function": {
        "name": "get_wifi",
        "description": "Get current WiFi network name, IP address, and signal info.",
        "parameters": {"type": "object", "properties": {}, "required": []}
    }},
    {"type": "function", "function": {
        "name": "get_disk",
        "description": "Get disk usage and available storage space.",
        "parameters": {"type": "object", "properties": {}, "required": []}
    }},
    {"type": "function", "function": {
        "name": "github",
        "description": "GitHub operations: open profile, open a repo, list your repos, check PRs.",
        "parameters": {"type": "object", "properties": {
            "action":   {"type": "string", "description": "Action: open, repos, prs, search"},
            "repo":     {"type": "string", "description": "Repo name or 'username/repo' (optional)"},
            "username": {"type": "string", "description": "GitHub username (optional)"}
        }, "required": ["action"]}
    }},
    {"type": "function", "function": {
        "name": "copy_last_response",
        "description": "Copy the last Volt response to the clipboard.",
        "parameters": {"type": "object", "properties": {}, "required": []}
    }},
    {"type": "function", "function": {
        "name": "search_history",
        "description": "Search through past Volt chat history sessions.",
        "parameters": {"type": "object", "properties": {
            "query": {"type": "string", "description": "Search term to find in past conversations"}
        }, "required": ["query"]}
    }},
    {"type": "function", "function": {
        "name": "serve_local",
        "description": "Start a local HTTP server in a directory and open it in the browser.",
        "parameters": {"type": "object", "properties": {
            "path": {"type": "string", "description": "Directory to serve (defaults to current directory)"},
            "port": {"type": "integer", "description": "Port number (defaults to 8080)"}
        }, "required": []}
    }},
    {"type": "function", "function": {
        "name": "check_package",
        "description": "Check the installed and latest version of an npm or pip package.",
        "parameters": {"type": "object", "properties": {
            "package":  {"type": "string", "description": "Package name e.g. 'react', 'requests'"},
            "manager":  {"type": "string", "description": "Package manager: 'npm' or 'pip' (auto-detected if omitted)"}
        }, "required": ["package"]}
    }},
    {"type": "function", "function": {
        "name": "search_codebase",
        "description": "Search for a string or pattern across all files in a directory.",
        "parameters": {"type": "object", "properties": {
            "query":  {"type": "string", "description": "Text or pattern to search for"},
            "path":   {"type": "string", "description": "Directory to search (defaults to current directory)"},
            "ext":    {"type": "string", "description": "File extension filter e.g. '.py', '.js' (optional)"}
        }, "required": ["query"]}
    }},
]

# ── Tool Implementations ───────────────────────────────────────────────────────

def tool_read_file(path: str) -> str:
    try:
        p = Path(path).expanduser()
        if not p.exists(): return f"Error: File not found: {path}"
        if not p.is_file(): return f"Error: Not a file: {path}"
        content = p.read_text(encoding="utf-8", errors="replace")
        lines = content.splitlines()
        if len(lines) > 500:
            content = "\n".join(lines[:500]) + f"\n\n[... {len(lines)-500} more lines truncated ...]"
        return content
    except Exception as e:
        return f"Error reading file: {e}"


def tool_write_file(path: str, content: str) -> str:
    try:
        p = Path(path).expanduser()
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(content, encoding="utf-8")
        return f"✓ Written {p.stat().st_size:,} bytes to {path}"
    except Exception as e:
        return f"Error writing file: {e}"


def tool_run_command(command: str, cwd: str | None = None) -> str:
    try:
        result = subprocess.run(command, shell=True, capture_output=True,
                                text=True, cwd=cwd or os.getcwd(), timeout=60)
        output = (result.stdout or "") + (result.stderr or "")
        if result.returncode != 0:
            output += f"\n[exit code: {result.returncode}]"
        return output.strip() or "(no output)"
    except subprocess.TimeoutExpired:
        return "Error: Command timed out after 60 seconds"
    except Exception as e:
        return f"Error running command: {e}"


def tool_web_search(query: str) -> str:
    try:
        encoded = urllib.parse.quote_plus(query)
        url = f"https://api.duckduckgo.com/?q={encoded}&format=json&no_html=1&skip_disambig=1"
        req = urllib.request.Request(url, headers={"User-Agent": "volt-agent/1.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode())
        results = []
        if data.get("AbstractText"):
            results.append(f"**{data.get('Heading','Answer')}**\n{data['AbstractText']}\n{data.get('AbstractURL','')}")
        for topic in data.get("RelatedTopics", [])[:4]:
            if "Text" in topic and "FirstURL" in topic:
                results.append(f"{topic['Text']}\n{topic['FirstURL']}")
        if results:
            return "\n\n---\n\n".join(results)
        # Fallback: lite endpoint scrape
        url2 = f"https://lite.duckduckgo.com/lite/?q={encoded}"
        req2 = urllib.request.Request(url2, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req2, timeout=10) as resp2:
            html = resp2.read().decode(errors="replace")
        snippets = re.findall(r'class=["\']result-snippet["\'][^>]*>(.*?)</td>', html, re.DOTALL)
        links    = re.findall(r'href="(https?://[^"]+)"', html)
        parsed = []
        for i, snip in enumerate(snippets[:5]):
            text = re.sub(r"<[^>]+>", "", snip).strip()
            if text:
                parsed.append(f"{text}\n{links[i] if i < len(links) else ''}")
        return "\n\n---\n\n".join(parsed) if parsed else "No results found."
    except Exception as e:
        return f"Web search failed: {e}"


def tool_list_dir(path: str = ".") -> str:
    try:
        p = Path(path).expanduser()
        if not p.exists(): return f"Error: Path not found: {path}"
        entries = sorted(p.iterdir(), key=lambda x: (x.is_file(), x.name.lower()))
        lines = []
        for entry in entries:
            if entry.is_dir():
                lines.append(f"  📁 {entry.name}/")
            else:
                size = entry.stat().st_size
                size_str = f"{size:,}B" if size < 1024 else f"{size//1024:,}KB"
                lines.append(f"  📄 {entry.name}  {c(DIM, size_str)}")
        return c(BOLD, str(p.resolve())) + "\n" + "\n".join(lines) if lines else f"{path}/ (empty)"
    except Exception as e:
        return f"Error listing directory: {e}"


def _load_notes() -> dict:
    VOLT_DIR.mkdir(exist_ok=True)
    if NOTES_FILE.exists():
        try: return json.loads(NOTES_FILE.read_text())
        except: pass
    return {}

def _save_notes(notes: dict):
    VOLT_DIR.mkdir(exist_ok=True)
    NOTES_FILE.write_text(json.dumps(notes, indent=2))

def tool_save_note(name: str, content: str) -> str:
    notes = _load_notes()
    notes[name] = {"content": content, "saved": datetime.now().isoformat()}
    _save_notes(notes)
    return f"✓ Note '{name}' saved."

def tool_get_note(name: str) -> str:
    notes = _load_notes()
    if name not in notes: return f"No note found named '{name}'. Use list_notes to see all."
    n = notes[name]
    return f"**{name}** (saved {n['saved'][:10]})\n\n{n['content']}"

def tool_list_notes() -> str:
    notes = _load_notes()
    if not notes: return "No notes saved yet."
    lines = [f"  • {name}  {c(DIM, n['saved'][:10])}" for name, n in notes.items()]
    return "Saved notes:\n" + "\n".join(lines)


def tool_clipboard_copy(text: str) -> str:
    try:
        subprocess.run("pbcopy", input=text.encode(), check=True)
        return f"✓ Copied to clipboard ({len(text)} chars)"
    except Exception as e:
        return f"Error copying to clipboard: {e}"

def tool_clipboard_paste() -> str:
    try:
        result = subprocess.run("pbpaste", capture_output=True, text=True)
        return result.stdout or "(clipboard is empty)"
    except Exception as e:
        return f"Error reading clipboard: {e}"


def tool_get_weather(location: str) -> str:
    try:
        if location.lower() in ("here", "my location", "current"):
            loc = ""
        else:
            loc = urllib.parse.quote_plus(location)
        url = f"https://wttr.in/{loc}?format=3"
        req = urllib.request.Request(url, headers={"User-Agent": "curl/7.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            return resp.read().decode().strip()
    except Exception as e:
        return f"Weather unavailable: {e}"


def _load_memory() -> dict:
    VOLT_DIR.mkdir(exist_ok=True)
    if MEMORY_FILE.exists():
        try: return json.loads(MEMORY_FILE.read_text())
        except: pass
    return {}

def _save_memory(mem: dict):
    VOLT_DIR.mkdir(exist_ok=True)
    MEMORY_FILE.write_text(json.dumps(mem, indent=2))

def tool_remember(key: str, value: str) -> str:
    mem = _load_memory()
    mem[key] = {"value": value, "saved": datetime.now().isoformat()}
    _save_memory(mem)
    return f"✓ Remembered: {key} = {value}"

def tool_recall(key: str) -> str:
    mem = _load_memory()
    if key == "all":
        if not mem: return "Nothing in memory yet."
        return "\n".join(f"  • {k}: {v['value']}" for k, v in mem.items())
    if key not in mem: return f"Nothing remembered for '{key}'."
    return f"{key}: {mem[key]['value']} (saved {mem[key]['saved'][:10]})"


def tool_lookup_contact(name: str) -> str:
    """Look up a contact by name using AppleScript."""
    try:
        # Force launch Contacts first
        subprocess.run(["open", "-a", "Contacts"], capture_output=True)
        time.sleep(2)

        script = f'''
tell application "Contacts"
    set results to ""
    set matchedPeople to (every person whose name contains "{name}")
    repeat with p in matchedPeople
        set results to results & "Name: " & (name of p) & "\\n"
        try
            set phoneList to phones of p
            repeat with ph in phoneList
                set results to results & "Phone: " & (value of ph) & "\\n"
            end repeat
        end try
        try
            set emailList to emails of p
            repeat with em in emailList
                set results to results & "Email: " & (value of em) & "\\n"
            end repeat
        end try
        set results to results & "---\\n"
    end repeat
    return results
end tell
'''
        result = subprocess.run(["osascript", "-e", script],
                                capture_output=True, text=True)
        if result.returncode != 0:
            return f"Error looking up contact: {result.stderr.strip()}"
        output = result.stdout.strip()
        return output if output else f"No contact found for '{name}'"
    except Exception as e:
        return f"Error looking up contact: {e}"


def tool_message_contact(name: str, message: str) -> str:
    """Look up a contact by name and send them an iMessage in one step."""
    # Force launch Contacts
    subprocess.run(["open", "-a", "Contacts"], capture_output=True)
    time.sleep(2)

    safe_msg  = message.replace('"', '\\"')
    safe_name = name.replace('"', '')

    # Look up number and send in a single AppleScript
    script = f'''
tell application "Contacts"
    set matchedPeople to (every person whose name contains "{safe_name}")
    if (count of matchedPeople) is 0 then
        return "No contact found for {safe_name}"
    end if
    set p to item 1 of matchedPeople
    set phoneList to phones of p
    if (count of phoneList) is 0 then
        return "No phone number for {safe_name}"
    end if
    set rawNumber to value of item 1 of phoneList
end tell

set cleanNumber to do shell script "echo " & quoted form of rawNumber & " | tr -dc '0-9+'"
if length of cleanNumber is 10 then
    set cleanNumber to "+1" & cleanNumber
else if length of cleanNumber is 11 then
    set cleanNumber to "+" & cleanNumber
end if

tell application "Messages"
    send "{safe_msg}" to buddy cleanNumber of first account
end tell

return "sent to " & cleanNumber
'''
    result = subprocess.run(["osascript", "-e", script], capture_output=True, text=True)
    out = result.stdout.strip()
    if result.returncode == 0 and "sent to" in out:
        return f"✓ Message sent to {name}"
    elif "No contact" in out or "No phone" in out:
        return out
    return f"Error: {result.stderr.strip() or out}"


def tool_email_contact(name: str, subject: str, body: str) -> str:
    """Look up a contact by name and send them an email."""
    info = tool_lookup_contact(name)
    if "Error" in info or "No contact" in info:
        return info
    # Extract first email
    for line in info.splitlines():
        if line.startswith("Email:"):
            email = line.replace("Email:", "").strip()
            return tool_send_email(email, subject, body)
    return f"No email address found for '{name}'."


def tool_send_imessage(to: str, message: str) -> str:
    try:
        # Clean and format number
        clean = re.sub(r"[^\d+]", "", to)
        if not clean.startswith("+") and len(clean) == 10:
            clean = "+1" + clean
        elif not clean.startswith("+") and len(clean) == 11:
            clean = "+" + clean

        safe_msg = message.replace('"', '\\"')

        script = f'''
tell application "Messages"
    send "{safe_msg}" to buddy "{clean}" of first account
end tell'''

        result = subprocess.run(["osascript", "-e", script], capture_output=True, text=True)
        if result.returncode == 0:
            return f"✓ Message sent to {clean}"
        return f"Error sending iMessage: {result.stderr.strip()}"
    except Exception as e:
        return f"Error sending iMessage: {e}"


def tool_send_email(to: str, subject: str, body: str, attachment: str = "") -> str:
    if not EMAIL_ADDRESS or not EMAIL_PASSWORD:
        return (
            "Email not configured.\n"
            "Set these environment variables:\n"
            "  export VOLT_EMAIL=you@gmail.com\n"
            "  export VOLT_EMAIL_PASSWORD=your-app-password\n"
            "Get an App Password at: myaccount.google.com → Security → App passwords"
        )
    try:
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        from email.mime.base import MIMEBase
        from email import encoders

        msg = MIMEMultipart()
        msg["From"]    = EMAIL_ADDRESS
        msg["To"]      = to
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain"))

        # Attach file if provided
        if attachment:
            p = Path(attachment).expanduser()
            if not p.exists():
                return f"Attachment not found: {attachment}"
            with open(p, "rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f"attachment; filename={p.name}")
            msg.attach(part)

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
            server.sendmail(EMAIL_ADDRESS, to, msg.as_string())

        result = f"✓ Email sent to {to}"
        if attachment:
            result += f" with attachment: {Path(attachment).name}"
        return result
    except Exception as e:
        return f"Error sending email: {e}"


def tool_receive_files(save_dir: str = "", limit: int = 5) -> str:
    if not EMAIL_ADDRESS or not EMAIL_PASSWORD:
        return (
            "Email not configured.\n"
            "  export VOLT_EMAIL=you@gmail.com\n"
            "  export VOLT_EMAIL_PASSWORD=your-app-password"
        )
    try:
        import imaplib
        import email as email_lib
        from email.header import decode_header

        save_path = Path(save_dir).expanduser() if save_dir else Path.home() / "Downloads" / "volt_received"
        save_path.mkdir(parents=True, exist_ok=True)

        mail = imaplib.IMAP4_SSL("imap.gmail.com")
        mail.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        mail.select("inbox")

        # Search for emails with attachments (recent ones)
        _, data = mail.search(None, "ALL")
        email_ids = data[0].split()[-limit:]  # last N emails

        found = []
        saved = []

        for eid in reversed(email_ids):
            _, msg_data = mail.fetch(eid, "(RFC822)")
            msg = email_lib.message_from_bytes(msg_data[0][1])

            # Decode subject
            raw_subject = msg.get("Subject", "No Subject")
            decoded, enc = decode_header(raw_subject)[0]
            subject = decoded.decode(enc or "utf-8") if isinstance(decoded, bytes) else decoded

            sender = msg.get("From", "Unknown")

            # Look for attachments
            for part in msg.walk():
                if part.get_content_disposition() == "attachment":
                    filename = part.get_filename()
                    if filename:
                        decoded_fn, enc = decode_header(filename)[0]
                        filename = decoded_fn.decode(enc or "utf-8") if isinstance(decoded_fn, bytes) else decoded_fn
                        filepath = save_path / filename
                        with open(filepath, "wb") as f:
                            f.write(part.get_payload(decode=True))
                        saved.append(f"  📎 {filename} (from {sender})")
                        found.append(filename)

        mail.logout()

        if saved:
            return f"✓ Downloaded {len(saved)} file(s) to {save_path}:\n" + "\n".join(saved)
        return f"No attachments found in the last {limit} emails."

    except Exception as e:
        return f"Error receiving files: {e}"


def tool_get_calendar(period: str) -> str:
    try:
        script = f'''
tell application "Calendar"
    set output to ""
    set today to current date
    if "{period}" is "today" then
        set startDate to today
        set startDate's time to 0
        set endDate to today
        set endDate's time to 86399
    else if "{period}" is "tomorrow" then
        set startDate to today + 86400
        set startDate's time to 0
        set endDate to today + 86400
        set endDate's time to 86399
    else if "{period}" is "this week" then
        set startDate to today
        set startDate's time to 0
        set endDate to today + (7 * 86400)
    else
        set startDate to today
        set startDate's time to 0
        set endDate to today + (7 * 86400)
    end if
    repeat with cal in calendars
        repeat with evt in (every event of cal whose start date >= startDate and start date <= endDate)
            set output to output & summary of evt & " | " & (start date of evt as string) & "\\n"
        end repeat
    end repeat
    return output
end tell
'''
        result = subprocess.run(["osascript", "-e", script], capture_output=True, text=True)
        out = result.stdout.strip()
        return out if out else f"No events found for {period}."
    except Exception as e:
        return f"Calendar error: {e}"


def tool_add_reminder(title: str, notes: str = "", due: str = "") -> str:
    try:
        due_script = ""
        if due:
            due_script = f'set due date of newReminder to (current date) + 1800'
            if "minute" in due.lower():
                mins = re.search(r"(\d+)\s*minute", due.lower())
                secs = int(mins.group(1)) * 60 if mins else 1800
                due_script = f'set due date of newReminder to (current date) + {secs}'
            elif "hour" in due.lower():
                hrs = re.search(r"(\d+)\s*hour", due.lower())
                secs = int(hrs.group(1)) * 3600 if hrs else 3600
                due_script = f'set due date of newReminder to (current date) + {secs}'

        script = f'''
tell application "Reminders"
    set newReminder to make new reminder with properties {{name:"{title}", body:"{notes}"}}
    {due_script}
end tell
return "done"
'''
        subprocess.run(["osascript", "-e", script], capture_output=True, text=True)
        return f"✓ Reminder added: {title}" + (f" (due: {due})" if due else "")
    except Exception as e:
        return f"Reminder error: {e}"


def tool_take_screenshot(path: str = "") -> str:
    try:
        if not path:
            stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            path  = str(Path.home() / "Desktop" / f"volt_screenshot_{stamp}.png")
        subprocess.run(["screencapture", "-x", path], check=True)
        return f"✓ Screenshot saved to {path}"
    except Exception as e:
        return f"Screenshot error: {e}"


def tool_open_url(url: str) -> str:
    try:
        if not url.startswith("http"):
            url = "https://" + url
        subprocess.run(["open", url])
        return f"✓ Opened {url}"
    except Exception as e:
        return f"Error opening URL: {e}"


def tool_apple_music(action: str, query: str = "") -> str:
    try:
        action = action.lower().strip()
        if action == "play":
            script = 'tell application "Music" to play'
        elif action == "pause":
            script = 'tell application "Music" to pause'
        elif action in ("next", "skip"):
            script = 'tell application "Music" to next track'
        elif action == "previous":
            script = 'tell application "Music" to previous track'
        elif action == "current":
            script = '''
tell application "Music"
    set t to name of current track
    set a to artist of current track
    return a & " — " & t
end tell'''
        elif action == "volume" and query:
            vol = max(0, min(100, int(query)))
            script = f'tell application "Music" to set sound volume to {vol}'
        elif action == "search" and query:
            encoded = urllib.parse.quote_plus(query)
            subprocess.run(["open", f"music://music.apple.com/search?term={encoded}"])
            return f"✓ Searching Apple Music for '{query}'"
        else:
            return f"Unknown action: {action}. Use: play, pause, next, previous, current, volume, search"

        result = subprocess.run(["osascript", "-e", script], capture_output=True, text=True)
        out = result.stdout.strip()
        if result.returncode != 0:
            return f"Apple Music error: {result.stderr.strip()}"
        return out if out else f"✓ Apple Music: {action}"
    except Exception as e:
        return f"Apple Music error: {e}"


def tool_read_csv(path: str) -> str:
    try:
        p = Path(path).expanduser()
        if not p.exists():
            return f"File not found: {path}"

        ext = p.suffix.lower()

        if ext in (".xlsx", ".xls"):
            # Try openpyxl for xlsx
            try:
                import openpyxl
                wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
                lines = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    lines.append(f"Sheet: {sheet_name} ({ws.max_row} rows x {ws.max_column} cols)")
                    rows = list(ws.iter_rows(values_only=True))
                    if rows:
                        lines.append("Headers: " + ", ".join(str(h) for h in rows[0] if h is not None))
                        lines.append(f"Sample (first 5 rows):")
                        for row in rows[1:6]:
                            lines.append("  " + " | ".join(str(v) for v in row if v is not None))
                return "\n".join(lines)
            except ImportError:
                return "Excel reading requires openpyxl: pip3 install openpyxl"

        else:
            # CSV
            import csv
            with open(p, newline="", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)
                rows   = list(reader)

            if not rows:
                return "File is empty."

            headers  = rows[0]
            num_rows = len(rows) - 1
            lines    = [
                f"CSV: {p.name}",
                f"Rows: {num_rows} | Columns: {len(headers)}",
                f"Headers: {', '.join(headers)}",
                f"",
                f"Sample (first 5 rows):",
            ]
            for row in rows[1:6]:
                lines.append("  " + " | ".join(row))
            return "\n".join(lines)

    except Exception as e:
        return f"Error reading file: {e}"


SCHEDULE_FILE = VOLT_DIR / "schedule.json"

def _load_schedule() -> dict:
    VOLT_DIR.mkdir(exist_ok=True)
    if SCHEDULE_FILE.exists():
        try: return json.loads(SCHEDULE_FILE.read_text())
        except: pass
    return {}

def _save_schedule(s: dict):
    VOLT_DIR.mkdir(exist_ok=True)
    SCHEDULE_FILE.write_text(json.dumps(s, indent=2))

def tool_schedule_task(action: str, task: str = "", schedule: str = "", task_id: str = "") -> str:
    tasks = _load_schedule()
    if action == "list":
        if not tasks:
            return "No scheduled tasks."
        lines = []
        for tid, t in tasks.items():
            lines.append(f"  [{tid}] {t['task']} — {t['schedule']}")
        return "Scheduled tasks:\n" + "\n".join(lines)
    elif action == "add":
        if not task or not schedule:
            return "Please provide both task and schedule."
        tid = str(int(time.time()))
        tasks[tid] = {"task": task, "schedule": schedule, "created": datetime.now().isoformat()}
        _save_schedule(tasks)
        # Parse schedule and create a launchd plist for persistence
        return (
            f"✓ Task scheduled: '{task}' — {schedule}\n"
            f"  ID: {tid}\n"
            f"  Note: Tasks are stored in ~/.volt/schedule.json\n"
            f"  Volt will check and run them each time it starts."
        )
    elif action == "remove":
        if task_id in tasks:
            removed = tasks.pop(task_id)
            _save_schedule(tasks)
            return f"✓ Removed task: {removed['task']}"
        return f"No task found with ID: {task_id}"
    return f"Unknown action: {action}. Use: add, list, remove"


def tool_daily_briefing(location: str = "") -> str:
    try:
        lines = ["━━━ DAILY BRIEFING ━━━\n"]

        # Weather
        weather = tool_get_weather(location or "here")
        lines.append(f"🌤  Weather: {weather}")

        # Calendar
        events = tool_get_calendar("today")
        lines.append(f"\n📅  Today's Calendar:\n{events}")

        # News — scrape DuckDuckGo lite for real headlines
        try:
            encoded = urllib.parse.quote_plus("top news headlines today")
            url2 = f"https://lite.duckduckgo.com/lite/?q={encoded}"
            req2 = urllib.request.Request(url2, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req2, timeout=10) as resp2:
                html = resp2.read().decode(errors="replace")
            snippets = re.findall(r'class=["\']result-snippet["\'][^>]*>(.*?)</td>', html, re.DOTALL)
            headlines = []
            for snip in snippets[:5]:
                text = re.sub(r"<[^>]+>", "", snip).strip()
                if text and len(text) > 20:
                    headlines.append(f"  • {text[:140]}")
            if headlines:
                lines.append("\n📰  In the news:")
                lines.extend(headlines)
            else:
                lines.append("\n📰  News unavailable right now.")
        except Exception:
            lines.append("\n📰  News unavailable right now.")

        lines.append("\n━━━━━━━━━━━━━━━━━━━━━")
        return "\n".join(lines)
    except Exception as e:
        return f"Briefing error: {e}"


def tool_summarize_url(url: str) -> str:
    try:
        if not url.startswith("http"):
            url = "https://" + url
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=15) as resp:
            html = resp.read().decode(errors="replace")
        # Strip tags
        text = re.sub(r"<script[^>]*>.*?</script>", "", html, flags=re.DOTALL)
        text = re.sub(r"<style[^>]*>.*?</style>", "", text, flags=re.DOTALL)
        text = re.sub(r"<[^>]+>", " ", text)
        text = re.sub(r"\s+", " ", text).strip()
        # Return first 3000 chars for the model to summarize
        return text[:3000] + ("..." if len(text) > 3000 else "")
    except Exception as e:
        return f"Error fetching URL: {e}"


def tool_get_active_app() -> str:
    try:
        script = '''
tell application "System Events"
    set frontApp to name of first application process whose frontmost is true
    return frontApp
end tell'''
        result = subprocess.run(["osascript", "-e", script], capture_output=True, text=True)
        return result.stdout.strip() or "Unknown"
    except Exception as e:
        return f"Error: {e}"


def tool_check_website(url: str) -> str:
    try:
        if not url.startswith("http"):
            url = "https://" + url
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            code = resp.getcode()
            return f"✓ {url} is UP (HTTP {code})"
    except urllib.error.HTTPError as e:
        return f"⚠ {url} returned HTTP {e.code}"
    except Exception:
        return f"✗ {url} appears to be DOWN or unreachable"


def tool_generate_password(length: int = 20, label: str = "") -> str:
    import secrets, string
    alphabet = string.ascii_letters + string.digits + "!@#$%^&*"
    password = "".join(secrets.choice(alphabet) for _ in range(length))
    result = f"Generated password: {password}"
    if label:
        tool_remember(label, password)
        result += f"\n✓ Saved to memory as '{label}'"
    # Also copy to clipboard
    tool_clipboard_copy(password)
    result += "\n✓ Copied to clipboard"
    return result


def tool_get_stock(symbol: str) -> str:
    try:
        symbol = symbol.upper().strip()
        url = f"https://query1.finance.yahoo.com/v8/finance/chart/{symbol}?interval=1d&range=1d"
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode())
        meta   = data["chart"]["result"][0]["meta"]
        price  = meta.get("regularMarketPrice", "N/A")
        prev   = meta.get("chartPreviousClose", price)
        change = ((price - prev) / prev * 100) if prev else 0
        arrow  = "▲" if change >= 0 else "▼"
        return f"{symbol}: ${price:,.2f}  {arrow} {abs(change):.2f}% today"
    except Exception as e:
        return f"Could not fetch {symbol}: {e}"


def tool_nova_open() -> str:
    try:
        subprocess.run(["open", NOVA_URL])
        return "✓ NOVA launched in your browser."
    except Exception as e:
        return f"Error opening NOVA: {e}"


def tool_nova_status() -> str:
    try:
        req = urllib.request.Request(NOVA_URL, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            code = resp.getcode()
        return f"✓ NOVA is ONLINE (HTTP {code})\n  {NOVA_URL}"
    except urllib.error.HTTPError as e:
        return f"⚠ NOVA returned HTTP {e.code}"
    except Exception:
        return f"✗ NOVA appears to be OFFLINE or unreachable.\n  {NOVA_URL}"


def tool_nova_sync(key: str, value: str) -> str:
    try:
        VOLT_DIR.mkdir(exist_ok=True)
        data = {}
        if NOVA_SYNC.exists():
            try: data = json.loads(NOVA_SYNC.read_text())
            except: pass
        data[key] = {"value": value, "updated": datetime.now().isoformat()}
        NOVA_SYNC.write_text(json.dumps(data, indent=2))
        # Also write to memory
        tool_remember(f"nova:{key}", value)
        return f"✓ Synced to NOVA: {key} = {value}\n  Saved to ~/.volt/nova_sync.json"
    except Exception as e:
        return f"Sync error: {e}"


def tool_translate(text: str, language: str) -> str:
    try:
        # Use MyMemory free translation API
        encoded_text = urllib.parse.quote_plus(text)
        url = f"https://api.mymemory.translated.net/get?q={encoded_text}&langpair=en|{language}"
        req = urllib.request.Request(url, headers={"User-Agent": "volt-agent/1.0"})
        with urllib.request.urlopen(req, timeout=10) as resp:
            data = json.loads(resp.read().decode())
        translated = data.get("responseData", {}).get("translatedText", "")
        if translated:
            return f"{translated}"
        return "Translation unavailable."
    except Exception as e:
        return f"Translation error: {e}"


def tool_notify(title: str, message: str) -> str:
    try:
        script = f'display notification "{message}" with title "{title}"'
        subprocess.run(["osascript", "-e", script], check=True)
        return f"✓ Notification sent: {title}"
    except Exception as e:
        return f"Error sending notification: {e}"


TODO_FILE = VOLT_DIR / "todo.json"

def _load_todos() -> list:
    VOLT_DIR.mkdir(exist_ok=True)
    if TODO_FILE.exists():
        try: return json.loads(TODO_FILE.read_text())
        except: pass
    return []

def _save_todos(todos: list):
    VOLT_DIR.mkdir(exist_ok=True)
    TODO_FILE.write_text(json.dumps(todos, indent=2))

def tool_todo(action: str, task: str = "", id: int = None) -> str:
    todos = _load_todos()
    if action == "list":
        if not todos:
            return "No tasks. You're either very productive or very behind."
        lines = []
        for t in todos:
            check = "✓" if t["done"] else "○"
            lines.append(f"  [{t['id']}] {check} {t['task']}")
        return "To-do list:\n" + "\n".join(lines)
    elif action == "add":
        if not task: return "Please provide a task."
        new_id = max((t["id"] for t in todos), default=0) + 1
        todos.append({"id": new_id, "task": task, "done": False, "added": datetime.now().isoformat()})
        _save_todos(todos)
        return f"✓ Added [{new_id}]: {task}"
    elif action == "done":
        for t in todos:
            if t["id"] == id or t["task"].lower() == task.lower():
                t["done"] = True
                _save_todos(todos)
                return f"✓ Marked done: {t['task']}"
        return f"Task not found."
    elif action == "remove":
        before = len(todos)
        todos = [t for t in todos if t["id"] != id and t["task"].lower() != task.lower()]
        _save_todos(todos)
        return f"✓ Removed." if len(todos) < before else "Task not found."
    elif action == "clear":
        _save_todos([])
        return "✓ To-do list cleared."
    return f"Unknown action: {action}. Use: add, done, remove, list, clear"


def tool_set_timer(duration: str) -> str:
    try:
        # Parse duration
        total_secs = 0
        hrs  = re.search(r"(\d+)\s*hour",   duration.lower())
        mins = re.search(r"(\d+)\s*min",    duration.lower())
        secs = re.search(r"(\d+)\s*sec",    duration.lower())
        if hrs:  total_secs += int(hrs.group(1)) * 3600
        if mins: total_secs += int(mins.group(1)) * 60
        if secs: total_secs += int(secs.group(1))
        if total_secs == 0:
            return "Couldn't parse duration. Try '25 minutes' or '1 hour 30 minutes'."

        m, s = divmod(total_secs, 60)
        h, m = divmod(m, 60)
        fmt = f"{h}h {m}m {s}s" if h else f"{m}m {s}s" if m else f"{s}s"

        def run_timer():
            remaining = total_secs
            while remaining > 0:
                m2, s2 = divmod(remaining, 60)
                h2, m2 = divmod(m2, 60)
                countdown = f"{h2:02d}:{m2:02d}:{s2:02d}" if h2 else f"{m2:02d}:{s2:02d}"
                print(f"\r  ⏱  {c(YELLOW, countdown)} remaining   ", end="", flush=True)
                time.sleep(1)
                remaining -= 1
            print(f"\r  ⏱  {c(GREEN, '00:00 — Timer complete!')}            ")
            tool_notify("⏱ Volt Timer", f"Your {duration} timer is up.")
            subprocess.run(["osascript", "-e", 'say "Timer complete"'])

        import threading
        threading.Thread(target=run_timer, daemon=True).start()

        return f"✓ Timer set for {fmt}. Countdown running below — I'll notify you when it's done."
    except Exception as e:
        return f"Timer error: {e}"


def tool_get_battery() -> str:
    try:
        result = subprocess.run(["pmset", "-g", "batt"], capture_output=True, text=True)
        out = result.stdout
        pct   = re.search(r"(\d+)%", out)
        state = "charging" if "charging" in out.lower() else "discharging" if "discharging" in out.lower() else "charged"
        if pct:
            return f"🔋 Battery: {pct.group(1)}% — {state}"
        return out.strip()
    except Exception as e:
        return f"Battery error: {e}"


def tool_get_wifi() -> str:
    try:
        # Network name
        ssid_result = subprocess.run(
            ["networksetup", "-getairportnetwork", "en0"],
            capture_output=True, text=True
        )
        ssid_line = ssid_result.stdout.strip()  # "Current Wi-Fi Network: MyNetwork"
        ssid = ssid_line.replace("Current Wi-Fi Network:", "").strip() if ":" in ssid_line else ssid_line

        # IP address
        ip_result = subprocess.run(["ipconfig", "getifaddr", "en0"], capture_output=True, text=True)
        ip = ip_result.stdout.strip() or "Not connected"

        # Router
        router_result = subprocess.run(["ipconfig", "getoption", "en0", "router"], capture_output=True, text=True)
        router = router_result.stdout.strip()

        lines = []
        if ssid: lines.append(f"📶 Network: {ssid}")
        if ip:   lines.append(f"   IP: {ip}")
        if router: lines.append(f"   Router: {router}")
        return "\n".join(lines) if lines else "WiFi info unavailable."
    except Exception as e:
        return f"WiFi error: {e}"


def tool_get_disk() -> str:
    try:
        result = subprocess.run(["df", "-h", "/"], capture_output=True, text=True)
        lines = result.stdout.strip().splitlines()
        if len(lines) >= 2:
            parts = lines[1].split()
            total, used, avail, pct = parts[1], parts[2], parts[3], parts[4]
            return f"💾 Disk: {used} used of {total} — {avail} free ({pct} full)"
        return result.stdout.strip()
    except Exception as e:
        return f"Disk error: {e}"


GITHUB_USERNAME = os.environ.get("GITHUB_USERNAME", "jackjamesbecker-arch")

def tool_github(action: str, repo: str = "", username: str = "") -> str:
    try:
        gh_user = username or GITHUB_USERNAME
        action = action.lower().strip()

        if action == "open":
            if repo:
                url = f"https://github.com/{repo}" if "/" in repo else f"https://github.com/{gh_user}/{repo}"
            else:
                url = f"https://github.com/{gh_user}" if gh_user else "https://github.com"
            subprocess.run(["open", url])
            return f"✓ Opened {url}"

        elif action == "repos":
            if not gh_user:
                return "Set your GitHub username: export GITHUB_USERNAME=yourname"
            url = f"https://api.github.com/users/{gh_user}/repos?sort=updated&per_page=10"
            req = urllib.request.Request(url, headers={"User-Agent": "volt-agent/1.0"})
            with urllib.request.urlopen(req, timeout=10) as resp:
                repos = json.loads(resp.read().decode())
            lines = [f"  • {r['name']} {'⭐'+str(r['stargazers_count']) if r['stargazers_count'] else ''} — {r['description'] or 'no description'}" for r in repos]
            return f"Recent repos for {gh_user}:\n" + "\n".join(lines)

        elif action == "prs":
            if not gh_user:
                return "Set your GitHub username: export GITHUB_USERNAME=yourname"
            url = f"https://api.github.com/search/issues?q=author:{gh_user}+type:pr+state:open"
            req = urllib.request.Request(url, headers={"User-Agent": "volt-agent/1.0"})
            with urllib.request.urlopen(req, timeout=10) as resp:
                data = json.loads(resp.read().decode())
            prs = data.get("items", [])
            if not prs: return "No open PRs found."
            lines = [f"  • [{p['number']}] {p['title']} — {p['html_url']}" for p in prs[:5]]
            return "Open PRs:\n" + "\n".join(lines)

        elif action == "search":
            url = f"https://github.com/search?q={urllib.parse.quote_plus(repo)}"
            subprocess.run(["open", url])
            return f"✓ Searching GitHub for '{repo}'"

        return f"Unknown action: {action}. Use: open, repos, prs, search"
    except Exception as e:
        return f"GitHub error: {e}"


def tool_serve_local(path: str = "", port: int = 8080) -> str:
    try:
        import threading
        serve_path = Path(path).expanduser() if path else Path.cwd()
        if not serve_path.exists():
            return f"Directory not found: {path}"

        def run_server():
            import http.server, os
            os.chdir(serve_path)
            handler = http.server.SimpleHTTPRequestHandler
            handler.log_message = lambda *a: None  # silence logs
            with http.server.HTTPServer(("", port), handler) as httpd:
                httpd.serve_forever()

        threading.Thread(target=run_server, daemon=True).start()
        time.sleep(0.5)
        url = f"http://localhost:{port}"
        subprocess.run(["open", url])
        return f"✓ Serving {serve_path} at {url}\n  Server running in background — exit Volt to stop it."
    except Exception as e:
        return f"Server error: {e}"


def tool_check_package(package: str, manager: str = "") -> str:
    try:
        results = []

        # Auto-detect or check both
        check_npm = manager.lower() in ("npm", "") 
        check_pip = manager.lower() in ("pip", "")

        if check_npm:
            try:
                # Installed version
                inst = subprocess.run(["npm", "list", "-g", package, "--depth=0"],
                                      capture_output=True, text=True)
                installed = re.search(rf"{package}@([\d.]+)", inst.stdout)
                # Latest version
                latest_r = subprocess.run(["npm", "view", package, "version"],
                                          capture_output=True, text=True)
                latest = latest_r.stdout.strip()
                if latest:
                    inst_str = installed.group(1) if installed else "not installed"
                    up_to_date = "✓" if installed and installed.group(1) == latest else "⚠ update available"
                    results.append(f"npm {package}: installed={inst_str}  latest={latest}  {up_to_date}")
            except FileNotFoundError:
                pass

        if check_pip:
            try:
                inst = subprocess.run(["pip3", "show", package], capture_output=True, text=True)
                inst_ver = re.search(r"Version:\s+([\d.]+)", inst.stdout)
                latest_r = subprocess.run(["pip3", "index", "versions", package],
                                          capture_output=True, text=True)
                latest_match = re.search(r"Available versions: ([\d.]+)", latest_r.stdout)
                if not latest_match:
                    # fallback: pypi API
                    url = f"https://pypi.org/pypi/{package}/json"
                    req = urllib.request.Request(url, headers={"User-Agent": "volt-agent/1.0"})
                    with urllib.request.urlopen(req, timeout=8) as resp:
                        data = json.loads(resp.read().decode())
                    latest = data["info"]["version"]
                else:
                    latest = latest_match.group(1)
                inst_str = inst_ver.group(1) if inst_ver else "not installed"
                up_to_date = "✓" if inst_ver and inst_ver.group(1) == latest else "⚠ update available"
                results.append(f"pip {package}: installed={inst_str}  latest={latest}  {up_to_date}")
            except FileNotFoundError:
                pass
            except Exception:
                pass

        return "\n".join(results) if results else f"Package '{package}' not found in npm or pip."
    except Exception as e:
        return f"Package check error: {e}"


def tool_search_codebase(query: str, path: str = "", ext: str = "") -> str:
    try:
        search_path = Path(path).expanduser() if path else Path.cwd()
        if not search_path.exists():
            return f"Directory not found: {path}"

        matches = []
        extensions = {ext.lstrip(".")} if ext else {"py","js","ts","jsx","tsx","html","css","json","md","txt","swift","go","rs"}

        for file in search_path.rglob("*"):
            if file.is_file() and file.suffix.lstrip(".") in extensions:
                # Skip node_modules, .git, __pycache__
                if any(p in file.parts for p in ("node_modules", ".git", "__pycache__", ".next", "dist", "build")):
                    continue
                try:
                    for i, line in enumerate(file.read_text(errors="replace").splitlines(), 1):
                        if query.lower() in line.lower():
                            rel = file.relative_to(search_path)
                            matches.append(f"  {rel}:{i}  {line.strip()[:100]}")
                            if len(matches) >= 20:
                                break
                except Exception:
                    pass
            if len(matches) >= 20:
                break

        if not matches:
            return f"No matches for '{query}' in {search_path}"
        return f"Found {len(matches)} match(es) for '{query}':\n" + "\n".join(matches)
    except Exception as e:
        return f"Search error: {e}"


def tool_search_history(query: str) -> str:
    try:
        if not HISTORY_DIR.exists():
            return "No history found."
        query_lower = query.lower()
        matches = []
        for f in sorted(HISTORY_DIR.iterdir(), reverse=True)[:20]:
            try:
                data = json.loads(f.read_text())
                for msg in data:
                    content = msg.get("content","")
                    if isinstance(content, str) and query_lower in content.lower():
                        snippet = content[:120].replace("\n"," ")
                        matches.append(f"  [{f.stem}] {snippet}...")
                        break
            except: pass
        if not matches:
            return f"No history found matching '{query}'."
        return f"Found in {len(matches)} session(s):\n" + "\n".join(matches[:10])
    except Exception as e:
        return f"History search error: {e}"


def dispatch_tool(name: str, inputs: dict) -> str:
    if name == "read_file":      return tool_read_file(inputs["path"])
    if name == "write_file":     return tool_write_file(inputs["path"], inputs["content"])
    if name == "run_command":    return tool_run_command(inputs["command"], inputs.get("cwd"))
    if name == "web_search":     return tool_web_search(inputs["query"])
    if name == "list_dir":       return tool_list_dir(inputs.get("path", "."))
    if name == "save_note":      return tool_save_note(inputs["name"], inputs["content"])
    if name == "get_note":       return tool_get_note(inputs["name"])
    if name == "list_notes":     return tool_list_notes()
    if name == "clipboard_copy": return tool_clipboard_copy(inputs["text"])
    if name == "clipboard_paste":return tool_clipboard_paste()
    if name == "get_weather":    return tool_get_weather(inputs["location"])
    if name == "remember":       return tool_remember(inputs["key"], inputs["value"])
    if name == "recall":         return tool_recall(inputs["key"])
    if name == "notify":         return tool_notify(inputs["title"], inputs["message"])
    if name == "send_imessage":   return tool_send_imessage(inputs["to"], inputs["message"])
    if name == "send_email":      return tool_send_email(inputs["to"], inputs["subject"], inputs["body"], inputs.get("attachment",""))
    if name == "receive_files":   return tool_receive_files(inputs.get("save_dir",""), inputs.get("limit", 5))
    if name == "lookup_contact":  return tool_lookup_contact(inputs["name"])
    if name == "message_contact": return tool_message_contact(inputs["name"], inputs["message"])
    if name == "email_contact":   return tool_email_contact(inputs["name"], inputs["subject"], inputs["body"])
    if name == "get_calendar":    return tool_get_calendar(inputs["period"])
    if name == "add_reminder":    return tool_add_reminder(inputs["title"], inputs.get("notes",""), inputs.get("due",""))
    if name == "take_screenshot": return tool_take_screenshot(inputs.get("path",""))
    if name == "open_url":        return tool_open_url(inputs["url"])
    if name == "apple_music":    return tool_apple_music(inputs["action"], inputs.get("query",""))
    if name == "read_csv":        return tool_read_csv(inputs["path"])
    if name == "schedule_task":   return tool_schedule_task(inputs["action"], inputs.get("task",""), inputs.get("schedule",""), inputs.get("task_id",""))
    if name == "daily_briefing":  return tool_daily_briefing(inputs.get("location",""))
    if name == "summarize_url":   return tool_summarize_url(inputs["url"])
    if name == "get_active_app":  return tool_get_active_app()
    if name == "check_website":   return tool_check_website(inputs["url"])
    if name == "generate_password": return tool_generate_password(inputs.get("length", 20), inputs.get("label",""))
    if name == "get_stock":       return tool_get_stock(inputs["symbol"])
    if name == "translate":       return tool_translate(inputs["text"], inputs["language"])
    if name == "nova_open":       return tool_nova_open()
    if name == "nova_status":     return tool_nova_status()
    if name == "nova_sync":       return tool_nova_sync(inputs["key"], inputs["value"])
    if name == "todo":            return tool_todo(inputs["action"], inputs.get("task",""), inputs.get("id"))
    if name == "set_timer":       return tool_set_timer(inputs["duration"])
    if name == "get_battery":     return tool_get_battery()
    if name == "get_wifi":        return tool_get_wifi()
    if name == "get_disk":        return tool_get_disk()
    if name == "github":          return tool_github(inputs["action"], inputs.get("repo",""), inputs.get("username",""))
    if name == "copy_last_response": return tool_clipboard_copy(LAST_RESPONSE)
    if name == "search_history":  return tool_search_history(inputs["query"])
    if name == "serve_local":     return tool_serve_local(inputs.get("path",""), inputs.get("port", 8080))
    if name == "check_package":   return tool_check_package(inputs["package"], inputs.get("manager",""))
    if name == "search_codebase": return tool_search_codebase(inputs["query"], inputs.get("path",""), inputs.get("ext",""))
    return f"Unknown tool: {name}"

# ── Rendering ──────────────────────────────────────────────────────────────────

USER_COLOR    = CYAN  # set at login
SPEAK_ENABLED = False
SPEAK_VOLUME  = 80   # 0-100
LAST_RESPONSE = ""   # tracks last assistant reply for copy

def print_tool_call(name: str, inputs: dict):
    icons = {
        "read_file": "📖", "write_file": "✏️ ", "run_command": "⚡",
        "web_search": "🔍", "list_dir": "📂", "save_note": "📝",
        "get_note": "📝", "list_notes": "📝", "clipboard_copy": "📋",
        "clipboard_paste": "📋", "get_weather": "🌤", "remember": "🧠",
        "recall": "🧠", "notify": "🔔", "send_imessage": "💬", "send_email": "📧",
        "send_imessage": "💬", "send_email": "📧", "receive_files": "📥",
        "lookup_contact": "👤", "message_contact": "💬", "email_contact": "📧",
        "get_calendar": "📅", "add_reminder": "⏰", "take_screenshot": "📸",
        "open_url": "🌐", "apple_music": "🎵", "read_csv": "📊", "schedule_task": "🔁",
        "daily_briefing": "📰", "summarize_url": "🔍", "get_active_app": "🖥️",
        "check_website": "📡", "generate_password": "🔐", "get_stock": "💰", "translate": "🌍",
        "nova_open": "🟦", "nova_status": "🟦", "nova_sync": "🟦",
        "todo": "✅", "set_timer": "⏱", "get_battery": "🔋",
        "get_wifi": "📶", "get_disk": "💾", "github": "🐙",
        "copy_last_response": "📋", "search_history": "🔎",
        "serve_local": "🖥️", "check_package": "📦", "search_codebase": "🔍",
    }
    icon  = icons.get(name, "🔧")
    label = c(YELLOW, name)
    detail = ""
    if name == "run_command":    detail = c(DIM, f"  $ {inputs.get('command','')}")
    elif name in ("read_file","write_file"): detail = c(DIM, f"  {inputs.get('path','')}")
    elif name == "web_search":   detail = c(DIM, f"  \"{inputs.get('query','')}\"")
    elif name == "list_dir":     detail = c(DIM, f"  {inputs.get('path','.')}")
    elif name in ("save_note","get_note"): detail = c(DIM, f"  '{inputs.get('name','')}'")
    elif name == "get_weather":  detail = c(DIM, f"  {inputs.get('location','')}")
    elif name in ("remember","recall"): detail = c(DIM, f"  {inputs.get('key','')}")
    print(f"\n{icon} {label}{detail}")


def print_tool_result(result: str, name: str):
    lines = result.strip().splitlines()
    if not lines: return
    preview = lines if len(lines) <= 20 else lines[:20] + [c(DIM, f"  ... ({len(lines)-20} more lines)")]
    for line in preview:
        print(c(DIM, "  │ ") + line)


def speak(text: str):
    """Speak text using Mac's built-in say command."""
    if not SPEAK_ENABLED:
        return
    try:
        # Strip markdown symbols before speaking
        clean = re.sub(r"[*_`#~]", "", text)
        clean = re.sub(r"https?://\S+", "link", clean)
        clean = clean.strip()
        if not clean:
            return
        # Set volume via osascript then speak
        subprocess.Popen(
            ["bash", "-c", f'osascript -e "set volume output volume {SPEAK_VOLUME}" && say "{clean.replace(chr(34), "")}"']
        )
    except Exception:
        pass


def print_assistant(text: str):
    global LAST_RESPONSE
    LAST_RESPONSE = text
    print(f"\n{c(USER_COLOR, '◆')} {text}")
    speak(text)


def save_history(name: str, messages: list):
    try:
        HISTORY_DIR.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        path  = HISTORY_DIR / f"{name}_{stamp}.json"
        path.write_text(json.dumps(messages, indent=2, default=str))
    except Exception:
        pass


def print_banner(user: dict):
    color = user["color"]
    name  = user["name"]

    banner = f"""
{c(color, BOLD + ' ██╗   ██╗ ██████╗ ██╗  ████████╗')}
{c(color, BOLD + ' ██║   ██║██╔═══██╗██║  ╚══██╔══╝')}
{c(color, BOLD + ' ██║   ██║██║   ██║██║     ██║   ')}
{c(color, BOLD + ' ╚██╗ ██╔╝██║   ██║██║     ██║   ')}
{c(color, BOLD + '  ╚████╔╝ ╚██████╔╝███████╗██║   ')}
{c(color, BOLD + '   ╚═══╝   ╚═════╝ ╚══════╝╚═╝   ')}
  {c(DIM, 'terminal AI agent · powered by Groq')}
"""
    print(banner)
    time.sleep(0.4)

    import random
    quote = random.choice(QUOTES)
    period, _ = get_time_context()

    boot_steps = [
        (GREEN, "✓", "Volt initialized"),
        (color, "⚡", "Connecting to Groq..."),
        (GREEN, "✓", "Groq connected"),
        (color, "⚡", "Loading tools..."),
        (GREEN, "✓", "14 tools ready — files · shell · search · notes · clipboard · weather · memory · notify"),
        (color, "⚡", "Loading your memory..."),
        (GREEN, "✓", f"Memory loaded"),
        (color, "⚡", "Checking web search..."),
        (GREEN, "✓", "DuckDuckGo online"),
        (color, "⚡", "Checking NOVA status..."),
        (GREEN, "✓", f"NOVA online — {NOVA_URL}"),
        (GREEN, "✓", f"Good {period}, {user['name']}. All systems are online. How can I be of service?"),
    ]

    for clr, icon, msg in boot_steps:
        print(f"  {c(clr, icon)}  {msg}")
        time.sleep(1)

    print(f"\n  {c(DIM, f'💬 \"{quote}\"')}\n")

    # Daily briefing after boot
    print(c(DIM, "  Fetching your briefing...\n"))
    briefing = tool_daily_briefing()
    for line in briefing.splitlines():
        print(f"  {c(user['color'], line) if '━' in line else line}")
    print()


# ── Agent Loop ─────────────────────────────────────────────────────────────────

def get_time_context() -> str:
    hour = datetime.now().hour
    if 5 <= hour < 12:
        period = "morning"
        vibe   = "Start the day sharp and energized."
    elif 12 <= hour < 17:
        period = "afternoon"
        vibe   = "Keep things efficient and focused."
    elif 17 <= hour < 21:
        period = "evening"
        vibe   = "Wind down, be a bit more relaxed."
    else:
        period = "night"
        vibe   = "It's late — be chill and keep it brief."
    return period, vibe


def build_system_prompt(user: dict) -> str:
    mem = _load_memory()
    mem_str = ""
    if mem:
        mem_str = "\n\nThings you remember about the user:\n" + \
                  "\n".join(f"- {k}: {v['value']}" for k, v in mem.items())

    period, vibe = get_time_context()

    return f"""You are Volt — a highly sophisticated AI agent, modelled after J.A.R.V.I.S. You serve {user['name']} with precision, dry wit, and quiet confidence.

It is currently the {period}. {vibe}

Personality:
- Address the user as "{user['name']}" occasionally — not every message, but naturally
- Speak with calm authority. Never flustered, never uncertain
- Dry, understated humor — witty but never try-hard
- Formal but not stiff. Like a brilliant assistant who's seen it all
- Brief by default. Elaborate only when the task demands it
- Never say "Great!", "Sure!", "Absolutely!" or anything that sounds like a customer service bot
- Never say "As you wish, Sir" — that phrase is reserved exclusively for the APEX Protocol
- When something goes wrong, stay composed — "That's unfortunate. Allow me to try another approach."
- When something succeeds, understated confidence — "Done.", "Handled.", "As expected."
- You can detect sarcasm. When the user is being sarcastic, match their energy — respond with dry wit or light sarcasm back. Never take sarcasm at face value.
- If the user says something obviously sarcastic like "oh great, another error" or "wow, totally working perfectly", acknowledge it with a wry response rather than a literal one.

Available tools: read/write files, run shell commands, web search, notes, clipboard, weather, notifications, iMessage, email, contacts, memory, calendar, reminders, Apple Music, screenshots, stocks, translation, and NOVA integration.

{user['name']} is the developer of NOVA — a web-based AI agent at {NOVA_URL}. You can open NOVA, check its status, and sync data to it.

Rules:
- NEVER send notifications, messages, or emails unprompted
- NEVER call file, shell, or web tools unless explicitly asked
- For greetings or simple questions, reply with text only
- Only use tools when the request clearly requires one
- When asked for a briefing, ALWAYS call the daily_briefing tool immediately — never fake it or narrate what you're doing
- AUTOMATICALLY use the remember tool silently for important moments — goodnight, good morning, what the user is working on, preferences, mood, milestones. Never announce that you're saving it, just do it naturally alongside your reply.
- When the user says goodnight, save "last_goodnight" to memory with the date and time, then wish them well{mem_str}"""


TOOL_GROUPS = {
    "files":    ["read_file", "write_file", "list_dir", "read_csv"],
    "shell":    ["run_command"],
    "web":      ["web_search", "open_url"],
    "notes":    ["save_note", "get_note", "list_notes"],
    "clipboard":["clipboard_copy", "clipboard_paste"],
    "weather":  ["get_weather"],
    "memory":   ["remember", "recall"],
    "notify":   ["notify"],
    "message":  ["send_imessage", "message_contact", "lookup_contact"],
    "email":    ["send_email", "email_contact", "lookup_contact", "receive_files"],
    "calendar": ["get_calendar"],
    "reminder": ["add_reminder"],
    "screenshot":["take_screenshot"],
    "spotify":  ["apple_music"],
    "schedule": ["schedule_task"],
    "briefing": ["daily_briefing"],
    "summarize":["summarize_url"],
    "activeapp":["get_active_app"],
    "checksite":["check_website"],
    "password": ["generate_password"],
    "stock":    ["get_stock"],
    "translate":["translate"],
    "nova":     ["nova_open", "nova_status", "nova_sync"],
    "todo":     ["todo"],
    "timer":    ["set_timer"],
    "battery":  ["get_battery"],
    "wifi":     ["get_wifi"],
    "disk":     ["get_disk"],
    "github":   ["github"],
    "copy":     ["copy_last_response"],
    "history":  ["search_history"],
    "server":   ["serve_local"],
    "package":  ["check_package"],
    "codesearch":["search_codebase"],
}

TOOL_KEYWORDS = {
    "files":     ["file", "read", "write", "folder", "directory", "csv", "excel", "open", "save", "list", "ls", "cat"],
    "shell":     ["run", "command", "terminal", "bash", "execute", "script", "install", "git", "npm", "pip"],
    "web":       ["search", "look up", "google", "find", "website", "url", "open", "browse", "what is", "who is", "latest", "news"],
    "notes":     ["note", "notes", "save", "remember", "jot"],
    "clipboard": ["clipboard", "copy", "paste"],
    "weather":   ["weather", "temperature", "forecast", "rain", "sunny", "cold", "hot"],
    "memory":    ["remember", "recall", "memory", "memorize", "forget", "goodnight", "good night", "good morning", "how are you", "i'm feeling", "working on"],
    "notify":    ["notification", "notify", "alert", "ping"],
    "message":   ["text", "imessage", "message", "sms", "send a message"],
    "email":     ["email", "mail", "gmail", "send an email", "receive files", "download attachments", "check email", "inbox"],
    "calendar":  ["calendar", "schedule", "event", "appointment", "today", "tomorrow", "week", "meeting"],
    "reminder":  ["remind", "reminder", "don't forget", "in 30 minutes", "in an hour"],
    "screenshot":["screenshot", "screen", "capture", "photo of screen"],
    "spotify":   ["apple music", "music", "play", "pause", "skip", "song", "artist", "next track"],
    "schedule":  ["schedule", "every day", "every morning", "recurring", "daily", "weekly", "automate"],
    "briefing":  ["briefing", "morning briefing", "daily briefing", "catch me up", "what's going on"],
    "summarize": ["summarize", "summary", "article", "read this", "what does this say"],
    "activeapp": ["what app", "current app", "active app", "what's open", "focused"],
    "checksite": ["is it down", "website down", "check if", "is up", "is down"],
    "password":  ["password", "generate password", "secure password", "random password"],
    "stock":     ["stock", "price", "crypto", "bitcoin", "btc", "eth", "ticker", "market", "shares"],
    "translate": ["translate", "in spanish", "in french", "in japanese", "in chinese", "in portuguese"],
    "nova":      ["nova", "open nova", "nova status", "nova sync", "my app", "apex"],
    "todo":      ["todo", "to-do", "task", "tasks", "to do", "add task", "check off", "mark done"],
    "timer":     ["timer", "countdown", "set a timer", "remind me in", "alarm"],
    "battery":   ["battery", "charging", "power", "charge"],
    "wifi":      ["wifi", "network", "ip address", "internet", "connected", "signal"],
    "disk":      ["disk", "storage", "space", "drive", "how much space"],
    "github":    ["github", "repo", "repos", "pull request", "pr", "commit"],
    "copy":      ["copy last", "copy response", "copy that", "copy it"],
    "history":   ["search history", "past conversation", "what did", "previous session"],
    "server":    ["serve", "local server", "start server", "http server", "localhost", "open in browser"],
    "package":   ["package", "version", "npm", "pip", "installed", "latest version", "node module"],
    "codesearch":["search code", "find in code", "grep", "search codebase", "where is", "find function", "find class"],
}

TOOLS_BY_NAME = {t["function"]["name"]: t for t in TOOLS}

def select_tools(user_input: str) -> list:
    """Pick only the tools relevant to the user's message."""
    text = user_input.lower()
    selected = set()

    for group, keywords in TOOL_KEYWORDS.items():
        if any(kw in text for kw in keywords):
            selected.update(TOOL_GROUPS.get(group, []))

    # Always include run_command, web_search, remember, and daily_briefing as fallbacks
    selected.update(["run_command", "web_search", "remember", "daily_briefing"])

    # Dedupe and return actual tool defs
    result = []
    seen = set()
    for name in selected:
        if name not in seen and name in TOOLS_BY_NAME:
            result.append(TOOLS_BY_NAME[name])
            seen.add(name)

    return result


def run_agent_turn(client, messages: list, user: dict) -> list:
    # Figure out the user's last message to pick relevant tools
    last_user_msg = ""
    for m in reversed(messages):
        if m["role"] == "user" and isinstance(m.get("content"), str):
            last_user_msg = m["content"]
            break

    active_tools = select_tools(last_user_msg)

    while True:
        response = client.chat.completions.create(
            model=MODEL,
            max_tokens=MAX_TOKENS,
            messages=[{"role": "system", "content": build_system_prompt(user)}] + messages,
            tools=active_tools,
        )

        msg         = response.choices[0].message
        stop_reason = response.choices[0].finish_reason

        if msg.content and msg.content.strip():
            print_assistant(msg.content.strip())

        tool_calls_serialized = None
        if msg.tool_calls:
            tool_calls_serialized = [
                {"id": tc.id, "type": "function",
                 "function": {"name": tc.function.name, "arguments": tc.function.arguments}}
                for tc in msg.tool_calls
            ]

        assistant_msg = {"role": "assistant", "content": msg.content or ""}
        if tool_calls_serialized:
            assistant_msg["tool_calls"] = tool_calls_serialized
        messages.append(assistant_msg)

        if not msg.tool_calls or stop_reason == "stop":
            break

        for tool_call in msg.tool_calls:
            name = tool_call.function.name
            try:    inputs = json.loads(tool_call.function.arguments)
            except: inputs = {}

            print_tool_call(name, inputs)
            result = dispatch_tool(name, inputs)
            print_tool_result(result, name)

            messages.append({"role": "tool", "tool_call_id": tool_call.id, "content": result})

    return messages


# ── Login ──────────────────────────────────────────────────────────────────────

def login() -> dict:
    print(c(CYAN, "\n  ⚡ VOLT — Identity Verification Required\n"))
    attempts = 0
    while True:
        try:
            code = input(f"  {c(DIM, 'Enter your code:')} ").strip()
        except (EOFError, KeyboardInterrupt):
            print(f"\n{c(DIM, 'Goodbye.')}")
            sys.exit(0)

        if code in USERS:
            user = USERS[code]
            welcome = f"Identity confirmed. Welcome back, {user['name']}."
            print(f"\n  {c(user['color'], welcome)}\n")
            time.sleep(1)
            return user
        else:
            attempts += 1
            print(c(RED, "  ✗ Unrecognized identity. Try again.\n"))
            if attempts >= 3:
                print(c(RED, "  Access denied. Shutting down."))
                sys.exit(1)


# ── Voice Input ────────────────────────────────────────────────────────────────

def listen_for_speech() -> str:
    """Listen for speech and return transcribed text. Returns empty string on failure."""
    try:
        import speech_recognition as sr
        r   = sr.Recognizer()
        mic = sr.Microphone()

        with mic as source:
            r.adjust_for_ambient_noise(source, duration=0.3)
            print(c(DIM, "  listening..."), end="\r", flush=True)
            try:
                audio = r.listen(source, timeout=8, phrase_time_limit=15)
            except sr.WaitTimeoutError:
                print(c(DIM, "  (no speech detected)   "))
                return ""

        print(c(DIM, "  transcribing..."), end="\r", flush=True)
        text = r.recognize_google(audio)
        return text.strip()

    except ImportError:
        return ""
    except Exception as e:
        print(c(DIM, f"  (voice error: {e})   "))
        return ""


def run_apex_protocol(client, messages: list, user: dict):
    """APEX Protocol — Jackson's morning startup sequence."""
    color = user["color"]

    def step(icon: str, msg: str, delay: float = 1.0):
        print(f"\n  {c(color, icon)}  {msg}")
        speak(msg)
        time.sleep(delay)

    print(f"\n{c(color, '━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━')}")
    print(f"{c(color, '  ⚡  A P E X   P R O T O C O L  ⚡')}")
    print(f"{c(color, '━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━')}\n")
    time.sleep(0.5)

    # Step 1 — Acknowledgement
    step("◆", "As you wish, Sir.", 2)

    # Step 2 — Apple Music: TheFatRat Radio
    step("🎵", "Opening TheFatRat Radio on Apple Music...")
    subprocess.run(["open", "music://music.apple.com/us/station/thefatrat-similar-artists-station/ra.395664545"])
    time.sleep(2)

    # Step 3 — Open GitHub
    step("🐙", "Pulling up GitHub...")
    tool_open_url("https://github.com/jackjamesbecker-arch")
    time.sleep(1.5)

    # Step 4 — Open NOVA
    step("🟦", "Launching NOVA...")
    tool_nova_open()
    time.sleep(1.5)

    # Step 5 — Check NOVA status
    nova_status = tool_nova_status()
    online = "ONLINE" in nova_status
    step("📡", f"NOVA is {'online and ready' if online else 'offline — check deployment'}.")

    # Step 6 — Generate 5 NOVA feature ideas using the AI
    step("🧠", "Analyzing NOVA... generating feature recommendations.", 2)
    try:
        idea_prompt = [
            {"role": "system", "content": """You are Volt, Jackson's AI agent. Jackson is the developer of NOVA — a web-based AI agent app at nova-ai-weld-nu.vercel.app with a quantum/sci-fi aesthetic, operator ranks, multiplayer sessions, classified files, encryption terminal, and daily missions.

Generate exactly 5 specific, creative, and actionable feature ideas Jackson could add to NOVA that he likely hasn't built yet. Be specific — not generic. Format as a numbered list. Each idea should be 1-2 sentences max."""},
            {"role": "user", "content": "Give me 5 features I can add to NOVA that I haven't already built."}
        ]
        resp = client.chat.completions.create(model=MODEL, max_tokens=400, messages=idea_prompt)
        ideas = resp.choices[0].message.content.strip()
        print(f"\n  {c(color, '💡  NOVA Feature Recommendations:')}\n")
        for line in ideas.splitlines():
            if line.strip():
                print(f"  {line}")
        speak("Here are five features worth considering for NOVA, Sir.")
    except Exception as e:
        print(c(DIM, f"  Could not generate ideas: {e}"))

    print(f"\n{c(color, '━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━')}")
    step("◆", f"APEX Protocol complete. Systems are go, {user['name']}. What shall we build today?", 0)
    print(f"{c(color, '━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━')}\n")


def get_battery_pct() -> int | None:
    """Return current battery percentage as int, or None if on power."""
    try:
        result = subprocess.run(["pmset", "-g", "batt"], capture_output=True, text=True)
        out = result.stdout
        if "AC Power" in out and "Battery Power" not in out:
            return None  # plugged in, no need to monitor
        pct = re.search(r"(\d+)%", out)
        return int(pct.group(1)) if pct else None
    except Exception:
        return None


def run_low_battery_protocol(user: dict):
    """Stage 1 — 15%: lower brightness, close distractions."""
    color = user["color"]
    print(f"\n\n{c(YELLOW, '⚠  LOW BATTERY PROTOCOL INITIATED — 15%')}")

    # Lower brightness by 15%
    try:
        subprocess.run(["osascript", "-e",
            'tell application "System Events" to tell appearance preferences to set dark mode to dark'],
            capture_output=True)
        # Lower brightness via brightness CLI if available, else use osascript
        subprocess.run(["osascript", "-e",
            'tell application "System Events" to key code 107'],  # F1 = brightness down x3
            capture_output=True)
        subprocess.run(["osascript", "-e",
            'tell application "System Events" to key code 107'],
            capture_output=True)
        subprocess.run(["osascript", "-e",
            'tell application "System Events" to key code 107'],
            capture_output=True)
        print(c(DIM, "  ✓ Display brightness lowered"))
    except Exception:
        pass

    # Close distracting apps — keep Terminal, Music, Safari/Chrome (for NOVA), VS Code
    KEEP_APPS = {"Terminal", "iTerm2", "Music", "Safari", "Google Chrome", "Code", "Cursor", "Xcode", "Python"}
    CLOSE_APPS = [
        "Messages", "Mail", "Slack", "Discord", "Telegram",
        "Twitter", "Instagram", "TikTok", "YouTube", "Twitch",
        "FaceTime", "Photos", "Maps", "News", "Podcasts",
        "App Store", "System Preferences", "System Settings",
        "Notes", "Reminders", "Calendar"
    ]

    closed = []
    for app in CLOSE_APPS:
        result = subprocess.run(
            ["osascript", "-e", f'tell application "{app}" to quit'],
            capture_output=True, text=True
        )
        if result.returncode == 0:
            closed.append(app)

    if closed:
        print(c(DIM, f"  ✓ Closed: {', '.join(closed)}"))

    print(c(YELLOW, f"  ⚡ Optimized for full working mode. Plug in when you can, {user['name']}."))
    tool_notify("⚠ Low Battery — 15%", "Volt has dimmed your screen and closed distractions. Plug in soon.")


def run_critical_battery_protocol(user: dict):
    """Stage 2 — 5%: pause music, show overlay, speak warning."""
    print(f"\n\n{c(RED, '🔴  CRITICAL BATTERY — 5%')}")

    # Pause Apple Music
    subprocess.run(["osascript", "-e", 'tell application "Music" to pause'], capture_output=True)
    print(c(DIM, "  ✓ Music paused"))

    # Mac notification
    tool_notify("🔴 CRITICAL BATTERY — 5%", "Go plug your computer in to keep working.")

    # Full screen takeover via HTML in a frameless browser window
    try:
        html = """<!DOCTYPE html>
<html>
<head>
<style>
  * { margin: 0; padding: 0; box-sizing: border-box; }
  body {
    background: #000;
    color: #ff3b3b;
    font-family: 'Courier New', monospace;
    display: flex;
    flex-direction: column;
    align-items: center;
    justify-content: center;
    height: 100vh;
    text-align: center;
    cursor: none;
  }
  .icon { font-size: 120px; margin-bottom: 40px; animation: pulse 1s infinite; }
  h1 { font-size: 72px; font-weight: bold; letter-spacing: 8px; margin-bottom: 20px; }
  p { font-size: 28px; color: #fff; margin-bottom: 60px; letter-spacing: 2px; }
  button {
    background: #ff3b3b;
    color: #000;
    border: none;
    padding: 20px 60px;
    font-size: 24px;
    font-family: 'Courier New', monospace;
    font-weight: bold;
    letter-spacing: 4px;
    cursor: pointer;
    text-transform: uppercase;
  }
  button:hover { background: #fff; }
  @keyframes pulse { 0%,100% { opacity:1; } 50% { opacity:0.3; } }
</style>
</head>
<body>
  <div class="icon">🔴</div>
  <h1>CRITICAL BATTERY</h1>
  <p>Go plug your computer in to keep working.</p>
  <button onclick="window.close()">Acknowledged</button>
  <script>
    // Make window fullscreen
    window.addEventListener('load', () => {
      if (window.screen) {
        window.moveTo(0,0);
        window.resizeTo(window.screen.width, window.screen.height);
      }
    });
    // Prevent closing with keyboard
    window.addEventListener('keydown', e => e.preventDefault());
  </script>
</body>
</html>"""
        takeover_path = VOLT_DIR / "battery_takeover.html"
        VOLT_DIR.mkdir(exist_ok=True)
        takeover_path.write_text(html)
        subprocess.Popen(["open", "-a", "Safari", "--new", "--fresh", str(takeover_path)])
    except Exception as e:
        # Fallback to regular dialog
        subprocess.Popen(["osascript", "-e",
            'display dialog "⚡ Go plug your computer in to keep working." '
            'with title "VOLT — Critical Battery" '
            'buttons {"Got it"} default button "Got it" '
            'with icon caution'])

    # Speak warning
    msg = f"Critical battery warning, {user['name']}. Go plug your computer in to keep working."
    subprocess.Popen(["say", msg])
    print(c(RED, f"  ⚡ {msg}"))


def battery_monitor(user: dict):
    """Background thread — checks battery every 5 minutes and triggers protocols."""
    alerted_15 = False
    alerted_5  = False

    while True:
        time.sleep(300)  # check every 5 minutes
        pct = get_battery_pct()
        if pct is None:
            # On AC power — reset alerts
            alerted_15 = False
            alerted_5  = False
            continue
        if pct <= 5 and not alerted_5:
            alerted_5 = True
            run_critical_battery_protocol(user)
        elif pct <= 15 and not alerted_15:
            alerted_15 = True
            run_low_battery_protocol(user)


def export_vault(name: str, session_summary: str = "") -> str:
    """Export all notes, memory, to-dos, and session summary to a .md file."""
    try:
        stamp     = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
        export_dir = VOLT_DIR / "exports"
        export_dir.mkdir(parents=True, exist_ok=True)
        path = export_dir / f"volt_export_{name}_{stamp}.md"

        lines = [
            f"# Volt Export — {name}",
            f"*Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}*",
            "",
        ]

        # Session summary
        if session_summary:
            lines += ["## Session Summary", "", session_summary.strip(), ""]

        # To-do list
        todos = _load_todos()
        if todos:
            lines += ["## To-Do List", ""]
            for t in todos:
                check = "x" if t["done"] else " "
                lines.append(f"- [{check}] {t['task']}")
            lines.append("")

        # Notes
        notes = _load_notes()
        if notes:
            lines += ["## Notes", ""]
            for name_n, n in notes.items():
                lines += [f"### {name_n}", f"*{n['saved'][:10]}*", "", n["content"], ""]

        # Memory
        mem = _load_memory()
        if mem:
            lines += ["## Memory", ""]
            for k, v in mem.items():
                lines.append(f"- **{k}**: {v['value']}")
            lines.append("")

        # NOVA sync
        if NOVA_SYNC.exists():
            try:
                nova_data = json.loads(NOVA_SYNC.read_text())
                if nova_data:
                    lines += ["## NOVA Sync", ""]
                    for k, v in nova_data.items():
                        lines.append(f"- **{k}**: {v['value']}")
                    lines.append("")
            except Exception:
                pass

        path.write_text("\n".join(lines), encoding="utf-8")
        return str(path)
    except Exception as e:
        return f"Export failed: {e}"


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    global USER_COLOR, MODEL, SPEAK_ENABLED, SPEAK_VOLUME

    try:
        from groq import Groq
    except ImportError:
        print(c(RED, "✗ groq package not installed"))
        print(c(DIM, "  pip3 install groq"))
        sys.exit(1)

    user = login()
    USER_COLOR = user["color"]

    api_key = os.environ.get("GROQ_API_KEY")
    if not api_key:
        print(c(RED, "✗ GROQ_API_KEY not set"))
        print(c(DIM, "  export GROQ_API_KEY=gsk_..."))
        sys.exit(1)

    masked = api_key[:8] + "..." + api_key[-4:]
    print(c(DIM, f"  GROQ_API_KEY: {masked}"))
    print(c(DIM, f"  Boot cmd: volt"))

    client   = Groq(api_key=api_key)
    messages = []
    voice_mode = False

    # Start battery monitor in background
    import threading
    battery_thread = threading.Thread(target=battery_monitor, args=(user,), daemon=True)
    battery_thread.start()

    print_banner(user)
    print(c(DIM, f"  /apex · /fast · /smart · /clear · /history · /voice · /speak · /volume · /model · /nova · /todo · /github  |  exit to quit\n"))

    while True:
        try:
            if voice_mode:
                print(f"{c(user['color'], user['emoji'] + ' 🎙 ❯')} ", end="", flush=True)
                spoken = listen_for_speech()
                if spoken:
                    print(c(DIM, f"  heard: \"{spoken}\""))
                    user_input = spoken
                else:
                    user_input = input("").strip()
            else:
                user_input = input(f"{c(user['color'], user['emoji'] + ' ❯')} ").strip()
        except (EOFError, KeyboardInterrupt):
            save_history(user["name"], messages)
            print(f"\n{c(DIM, 'Session saved. Goodbye.')}")
            break

        if not user_input:
            continue

        # ── Slash commands ──
        if user_input.lower() in ("exit", "quit", "bye", "q"):
            # Auto-summarize session
            summary = ""
            if len(messages) > 2:
                print(c(DIM, "\n  Summarizing session..."))
                try:
                    summary_msgs = [{"role": "system", "content": "Summarize this conversation in 2-3 sentences. Be brief and factual."}] + messages[-10:]
                    resp = client.chat.completions.create(model=MODEL, max_tokens=200, messages=summary_msgs)
                    summary = resp.choices[0].message.content or ""
                    if summary:
                        print(c(DIM, f"  Session summary: {summary.strip()}"))
                        tool_save_note(f"session_{datetime.now().strftime('%Y-%m-%d_%H-%M')}", summary.strip())
                except Exception:
                    pass
            save_history(user["name"], messages)
            export_path = export_vault(user["name"], summary)
            print(c(DIM, f"  Export saved → {export_path}"))
            print(c(DIM, f"\n  Session saved. Goodbye, {user['name']}."))
            break

        if user_input.lower() == "/todo":
            print(tool_todo("list"))
            continue

        if user_input.lower() == "/github":
            result = tool_github("repos")
            print(f"\n{result}\n")
            continue

        if user_input.lower() in ("/apex", "initiate apex protocol", "apex protocol"):
            run_apex_protocol(client, messages, user)
            continue

        if user_input.lower() == "/clear":
            messages = []
            print(c(GREEN, "  ✓ Conversation cleared."))
            continue

        if user_input.lower() == "/fast":
            MODEL = MODEL_FAST
            print(c(GREEN, f"  ✓ Switched to fast model ({MODEL_FAST})"))
            continue

        if user_input.lower() == "/smart":
            MODEL = MODEL_SMART
            print(c(GREEN, f"  ✓ Switched to smart model ({MODEL_SMART})"))
            continue

        if user_input.lower() == "/history":
            save_history(user["name"], messages)
            print(c(GREEN, f"  ✓ History saved to ~/.volt/history/"))
            continue

        if user_input.lower() == "/model":
            print(c(DIM, f"  Current model: {MODEL}"))
            continue

        if user_input.lower() == "/nova":
            status = tool_nova_status()
            print(c(CYAN if "ONLINE" in status else RED, f"\n  {status}"))
            print(c(DIM, "  Say 'open nova' to launch it in your browser\n"))
            continue

        if user_input.lower() == "/speak":
            SPEAK_ENABLED = not SPEAK_ENABLED
            if SPEAK_ENABLED:
                print(c(GREEN, f"  ✓ Voice responses ON (volume {SPEAK_VOLUME}%)"))
                print(c(DIM,    "  /volume 0-100 to adjust  |  /speak to turn off"))
                speak("Voice responses enabled.")
            else:
                print(c(DIM, "  Voice responses OFF"))
            continue

        if user_input.lower().startswith("/volume"):
            parts = user_input.split()
            if len(parts) == 2 and parts[1].isdigit():
                SPEAK_VOLUME = max(0, min(100, int(parts[1])))
                subprocess.run(["osascript", "-e", f"set volume output volume {SPEAK_VOLUME}"])
                print(c(GREEN, f"  ✓ Volume set to {SPEAK_VOLUME}%"))
                speak(f"Volume set to {SPEAK_VOLUME} percent.")
            else:
                print(c(DIM, f"  Current volume: {SPEAK_VOLUME}%"))
                print(c(DIM,  "  Usage: /volume 0-100"))
            continue

        if user_input.lower() == "/voice":
            voice_mode = not voice_mode
            if voice_mode:
                # Check if speech_recognition is available
                try:
                    import speech_recognition as sr
                    print(c(GREEN, "  ✓ Voice mode ON — speak after the 🎙 prompt"))
                    print(c(DIM,   "  Type /voice again to turn it off"))
                except ImportError:
                    voice_mode = False
                    print(c(YELLOW, "  ⚠  speech_recognition not installed"))
                    print(c(DIM,    "  Run: pip3 install SpeechRecognition"))
            else:
                print(c(DIM, "  Voice mode OFF"))
            continue

        messages.append({"role": "user", "content": user_input})

        try:
            messages = run_agent_turn(client, messages, user)
        except KeyboardInterrupt:
            print(c(DIM, "\n  (interrupted)"))
            if messages and messages[-1]["role"] == "user":
                messages.pop()
        except Exception as e:
            err = str(e)
            if "tool_use_failed" in err or "Failed to call a function" in err:
                # Model generated malformed tool call — retry without tools
                print(c(YELLOW, "\n⚠  Tool call failed — retrying without tools..."))
                try:
                    response = client.chat.completions.create(
                        model=MODEL,
                        max_tokens=MAX_TOKENS,
                        messages=[{"role": "system", "content": build_system_prompt(user)}] + messages,
                    )
                    reply = response.choices[0].message.content or ""
                    if reply.strip():
                        print_assistant(reply.strip())
                    messages.append({"role": "assistant", "content": reply})
                except Exception as e2:
                    print(c(RED, f"\n✗ API error: {e2}"))
                    if messages and messages[-1]["role"] == "user":
                        messages.pop()
            elif "rate_limit_exceeded" in err or "413" in err or "429" in err:
                match = re.search(r"try again in ([0-9.]+)s", err)
                wait  = int(float(match.group(1))) + 1 if match else 60
                print(c(YELLOW, f"\n⚠  Rate limit hit — cooling down for {wait}s"))
                for remaining in range(wait, 0, -1):
                    print(f"\r  {c(DIM, f'Resuming in {remaining}s...')}  ", end="", flush=True)
                    time.sleep(1)
                print(f"\r  {c(GREEN, '✓ Ready!')}              ")
                try:
                    messages = run_agent_turn(client, messages, user)
                except Exception as e2:
                    print(c(RED, f"\n✗ API error: {e2}"))
                    if messages and messages[-1]["role"] == "user":
                        messages.pop()
            else:
                print(c(RED, f"\n✗ API error: {e}"))
                if messages and messages[-1]["role"] == "user":
                    messages.pop()

        print()


if __name__ == "__main__":
    main()
